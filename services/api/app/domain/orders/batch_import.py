from __future__ import annotations

import csv
from dataclasses import dataclass, field
from io import StringIO
from pathlib import Path
from typing import Any, Protocol
import warnings
from uuid import uuid4

from app.domain import DomainError
from app.domain.orders.issues import ReviewIssue


DEFAULT_LISTING_ID = "birth-flower-card"
ORDER_STATUSES = {
    "IMPORTED",
    "READY",
    "NEEDS_REVIEW",
    "BLOCKED",
    "FAILED",
}


@dataclass
class BatchOrderItem:
    order_job_id: str
    batch_id: str
    row_number: int
    status: str
    order_id: str
    listing_id: str
    order_note: str
    source_adapter: str
    personalization: str = ""
    variation: str = ""
    listing_version: str | None = None
    raw_row: dict[str, str] = field(default_factory=dict)
    issues: list[ReviewIssue] = field(default_factory=list)
    customer_name: str | None = None
    month: int | None = None
    flower: str | None = None
    color: str | None = None
    font_option_no: int | None = None
    font_id: str | None = None
    parsed_order: Any | None = None


@dataclass
class BatchImport:
    batch_id: str
    source_name: str
    source_adapter: str
    items: list[BatchOrderItem]


class OrderSourceAdapter(Protocol):
    name: str

    def load(
        self,
        source: Path,
        *,
        batch_id: str,
        default_listing_id: str,
    ) -> BatchImport:
        ...


def import_orders(
    source: Path | str,
    *,
    adapter_name: str | None = None,
    batch_id: str | None = None,
    default_listing_id: str = DEFAULT_LISTING_ID,
) -> BatchImport:
    source_path = Path(source)
    if not source_path.is_file():
        raise DomainError(
            code="ORDER_SOURCE_NOT_FOUND",
            message="Order source file was not found.",
            details={"source": str(source_path)},
            recoverable=True,
        )
    adapter = _adapter_for_source(source_path, adapter_name)
    resolved_batch_id = batch_id or f"batch_{uuid4().hex}"
    return adapter.load(source_path, batch_id=resolved_batch_id, default_listing_id=default_listing_id)


def import_batch_csv(
    csv_content: str,
    *,
    source_name: str = "orders.csv",
    batch_id: str | None = None,
    default_listing_id: str = DEFAULT_LISTING_ID,
) -> BatchImport:
    resolved_batch_id = batch_id or f"batch_{uuid4().hex}"
    return _load_generic_csv_content(
        csv_content,
        source_name=source_name,
        batch_id=resolved_batch_id,
        default_listing_id=default_listing_id,
    )


def _adapter_for_source(source: Path, adapter_name: str | None) -> OrderSourceAdapter:
    adapters: dict[str, OrderSourceAdapter] = {
        "dianxiaomi-xlsx": DianxiaomiXlsxAdapter(),
        "generic-csv": GenericCsvAdapter(),
    }
    if adapter_name:
        try:
            return adapters[adapter_name]
        except KeyError as exc:
            raise DomainError(
                code="ORDER_ADAPTER_UNSUPPORTED",
                message="Order source adapter is not supported.",
                details={"adapter": adapter_name},
                recoverable=True,
            ) from exc

    suffix = source.suffix.casefold()
    if suffix == ".xlsx":
        return adapters["dianxiaomi-xlsx"]
    if suffix == ".csv":
        return adapters["generic-csv"]
    raise DomainError(
        code="ORDER_ADAPTER_UNSUPPORTED",
        message="Order source extension is not supported.",
        details={"extension": suffix},
        recoverable=True,
    )


class DianxiaomiXlsxAdapter:
    name = "dianxiaomi-xlsx"

    def load(
        self,
        source: Path,
        *,
        batch_id: str,
        default_listing_id: str,
    ) -> BatchImport:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DomainError(
                code="DEPENDENCY_MISSING",
                message="openpyxl is required to import Dianxiaomi XLSX files.",
                details={"package": "openpyxl"},
                recoverable=False,
            ) from exc

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Workbook contains no default style")
            workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        items: list[BatchOrderItem] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            order_id = _clean_cell(row[0] if len(row) > 0 else "")
            order_note = _clean_cell(row[1] if len(row) > 1 else "")
            if not order_id and not order_note:
                continue
            items.append(
                _build_item(
                    batch_id=batch_id,
                    row_number=row_number,
                    source_adapter=self.name,
                    order_id=order_id,
                    listing_id=default_listing_id,
                    order_note=order_note,
                    personalization="",
                    variation="",
                    raw_row={
                        "orderId": order_id,
                        "orderNote": order_note,
                    },
                )
            )
        workbook.close()
        return BatchImport(
            batch_id=batch_id,
            source_name=source.name,
            source_adapter=self.name,
            items=items,
        )


class GenericCsvAdapter:
    name = "generic-csv"

    def load(
        self,
        source: Path,
        *,
        batch_id: str,
        default_listing_id: str,
    ) -> BatchImport:
        return _load_generic_csv_content(
            source.read_text(encoding="utf-8-sig"),
            source_name=source.name,
            batch_id=batch_id,
            default_listing_id=default_listing_id,
        )


def _build_item(
    *,
    batch_id: str,
    row_number: int,
    source_adapter: str,
    order_id: str,
    listing_id: str,
    order_note: str,
    personalization: str,
    variation: str,
    listing_version: str | None = None,
    raw_row: dict[str, str] | None = None,
) -> BatchOrderItem:
    issues = _required_value_issues(order_id, order_note)
    return BatchOrderItem(
        order_job_id=f"job_{uuid4().hex}",
        batch_id=batch_id,
        row_number=row_number,
        status="BLOCKED" if issues else "IMPORTED",
        order_id=order_id,
        listing_id=listing_id,
        listing_version=listing_version,
        order_note=order_note,
        source_adapter=source_adapter,
        personalization=personalization,
        variation=variation,
        raw_row=raw_row or {},
        issues=issues,
    )


def _load_generic_csv_content(
    content: str,
    *,
    source_name: str,
    batch_id: str,
    default_listing_id: str,
) -> BatchImport:
    reader = csv.DictReader(StringIO(content.lstrip("\ufeff")))
    fieldnames = list(reader.fieldnames or [])
    missing_columns = [column for column in ("orderId", "orderNote") if column not in fieldnames]
    if missing_columns:
        raise DomainError(
            code="CSV_INVALID",
            message="CSV is missing required columns.",
            details={"missingColumns": missing_columns},
            recoverable=True,
        )

    items: list[BatchOrderItem] = []
    for row_number, row in enumerate(reader, start=2):
        normalized = {key: _clean_cell(value) for key, value in row.items() if key is not None}
        extra_values = row.get(None)
        item = _build_item(
            batch_id=batch_id,
            row_number=row_number,
            source_adapter=GenericCsvAdapter.name,
            order_id=normalized.get("orderId", ""),
            listing_id=normalized.get("listingId", "") or default_listing_id,
            order_note=normalized.get("orderNote", ""),
            personalization=normalized.get("personalization", ""),
            variation=normalized.get("variation", ""),
            listing_version=normalized.get("listingVersion") or None,
            raw_row=normalized,
        )
        if extra_values is not None:
            item.issues.append(
                ReviewIssue(
                    code="CSV_ROW_INVALID",
                    severity="blocking",
                    field=None,
                    message="CSV row has unexpected extra columns.",
                    raw_value=", ".join(_clean_cell(value) for value in extra_values),
                    requires_manual_action=True,
                )
            )
            item.status = "BLOCKED"
        items.append(item)
    return BatchImport(
        batch_id=batch_id,
        source_name=source_name,
        source_adapter=GenericCsvAdapter.name,
        items=items,
    )


def _required_value_issues(order_id: str, order_note: str) -> list[ReviewIssue]:
    issues: list[ReviewIssue] = []
    if not order_id:
        issues.append(_row_issue("orderId", "Order id is required."))
    if not order_note:
        issues.append(_row_issue("orderNote", "Order note is required."))
    return issues


def _row_issue(field: str, message: str) -> ReviewIssue:
    return ReviewIssue(
        code="CSV_ROW_INVALID",
        severity="blocking",
        field=field,
        message=message,
        requires_manual_action=True,
    )


def _clean_cell(value: object) -> str:
    return str(value or "").strip()
