from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def test_export_pool_writes_dianxiaomi_xlsx(client):
    client.post("/inbox/orders", json={"schema_version": "1.0", "order_id": "ORD-1", "remark": "first remark"})
    client.post("/inbox/orders", json={"schema_version": "1.0", "order_id": "ORD-2", "remark": "second remark"})

    body = client.post("/inbox/batch/export").json()
    assert body["count"] == 2
    assert body["path"] is not None

    workbook = load_workbook(Path(body["path"]), read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    # 首行表头（dianxiaomi 适配器会跳过），数据从第 2 行起：A=order_id, B=remark。
    assert rows[0] == ("订单号", "备注")
    assert {row[0]: row[1] for row in rows[1:]} == {"ORD-1": "first remark", "ORD-2": "second remark"}

    # 导出后订单标记为 QUEUED_FOR_BATCH。
    assert client.get("/inbox/orders/ORD-1").json()["status"] == "QUEUED_FOR_BATCH"


def test_export_pool_empty_returns_zero(client):
    body = client.post("/inbox/batch/export").json()
    assert body["count"] == 0
    assert body["path"] is None
