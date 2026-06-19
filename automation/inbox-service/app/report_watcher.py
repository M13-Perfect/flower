from __future__ import annotations

import threading
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session, sessionmaker

from app.db import session_scope
from app.models import STATUS_CANNOT_AUTOGEN, STATUS_DONE, Order, utcnow

# 对齐 services/api report.py 的列名。
COL_ORDER_ID = "订单号"
COL_STATUS = "状态"
COL_NEEDS_MANUAL = "是否需人工核验"
COL_REASON = "原因汇总"

EXPORTED_STATUS = "EXPORTED"
MANUAL_STATUSES = {"BLOCKED", "NEEDS_REVIEW", "FAILED"}


def read_report_rows(report_path: Path) -> list[dict]:
    """读 Flower 批量报告 xlsx；按表头名定位列，返回每行 {order_id,status,needs_manual,reason}。"""
    workbook = load_workbook(report_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    index = {name: i for i, name in enumerate(header)}

    def cell(row: tuple, name: str) -> str:
        i = index.get(name)
        if i is None or i >= len(row):
            return ""
        value = row[i]
        return "" if value is None else str(value).strip()

    result: list[dict] = []
    for row in rows[1:]:
        if row is None:
            continue
        order_id = cell(row, COL_ORDER_ID)
        if not order_id:
            continue
        result.append(
            {
                "order_id": order_id,
                "status": cell(row, COL_STATUS),
                "needs_manual": cell(row, COL_NEEDS_MANUAL) == "是",
                "reason": cell(row, COL_REASON),
            }
        )
    return result


def apply_report(session: Session, rows: list[dict]) -> dict:
    """把报告行回写到订单：EXPORTED→已完成；需人工核验/BLOCKED/NEEDS_REVIEW/FAILED→无法自动生成+原因。"""
    done = cannot = unknown = skipped = 0
    for row in rows:
        order = session.get(Order, row["order_id"])
        if order is None:
            skipped += 1  # 报告里有但本池没有的订单（如手工导入的 xlsx）——忽略。
            continue
        status_text = row["status"]
        if status_text == EXPORTED_STATUS and not row["needs_manual"]:
            order.status = STATUS_DONE
            order.reason = None
            order.done_at = utcnow()
            done += 1
        elif row["needs_manual"] or status_text in MANUAL_STATUSES:
            order.status = STATUS_CANNOT_AUTOGEN
            order.reason = row["reason"] or status_text
            cannot += 1
        else:
            unknown += 1
    return {"done": done, "cannot_autogen": cannot, "unknown": unknown, "skipped": skipped}


def apply_report_file(factory: sessionmaker[Session], report_path: Path) -> dict:
    rows = read_report_rows(Path(report_path))
    with session_scope(factory) as session:
        return apply_report(session, rows)


class ReportWatcher:
    """轮询 reports 目录里的 *-report.xlsx，新出现/被改动的就回写订单状态（按 mtime 去重）。"""

    def __init__(self, factory: sessionmaker[Session], reports_dir: Path) -> None:
        self._factory = factory
        self._reports_dir = Path(reports_dir)
        self._seen: dict[str, float] = {}
        self._stop = threading.Event()

    def scan_once(self) -> list[Path]:
        if not self._reports_dir.is_dir():
            return []
        applied: list[Path] = []
        for path in sorted(self._reports_dir.glob("*-report.xlsx")):
            try:
                mtime = path.stat().st_mtime
            except OSError:
                continue
            if self._seen.get(path.name) == mtime:
                continue
            try:
                apply_report_file(self._factory, path)
            except Exception:
                continue  # 单个报告读失败不影响其它；下次 mtime 变化会重试
            self._seen[path.name] = mtime
            applied.append(path)
        return applied

    def run_forever(self, interval: float = 3.0) -> None:
        while not self._stop.wait(interval):
            self.scan_once()

    def stop(self) -> None:
        self._stop.set()
